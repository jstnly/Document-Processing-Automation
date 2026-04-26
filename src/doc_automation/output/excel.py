"""Excel (.xlsx) output adapter using openpyxl."""

from __future__ import annotations

import logging
from pathlib import Path

from doc_automation.extraction.invoice import Invoice
from doc_automation.output.base import OutputAdapter

logger = logging.getLogger(__name__)


class ExcelAdapter(OutputAdapter):
    """Append rows to an Excel workbook, creating it if it doesn't exist."""

    def __init__(
        self,
        file_path: Path,
        columns: list[str],
        *,
        sheet_name: str = "Invoices",
        append: bool = True,
    ) -> None:
        self.file_path = file_path
        self.columns = columns
        self.sheet_name = sheet_name
        self.append = append
        self.file_path.parent.mkdir(parents=True, exist_ok=True)

    def write_rows(self, invoices: list[Invoice]) -> int:
        import openpyxl

        if self.append and self.file_path.exists():
            wb = openpyxl.load_workbook(str(self.file_path))
            ws = wb[self.sheet_name] if self.sheet_name in wb.sheetnames else wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.sheet_name  # type: ignore[union-attr]
            ws.append(self.columns)  # type: ignore[union-attr]

        for invoice in invoices:
            row_dict = invoice.to_dict()
            ws.append([row_dict.get(col, "") for col in self.columns])  # type: ignore[union-attr]

        wb.save(str(self.file_path))
        logger.info("Excel: wrote %d rows to %s", len(invoices), self.file_path)
        return len(invoices)
