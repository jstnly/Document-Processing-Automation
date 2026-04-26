"""Tests for output adapters: CSV, Excel, and Google Sheets (mocked)."""

from __future__ import annotations

import csv
from decimal import Decimal
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

from doc_automation.extraction.invoice import Invoice
from doc_automation.output import build_adapter
from doc_automation.output.csv_writer import CSVAdapter
from doc_automation.output.excel import ExcelAdapter
from doc_automation.output.sheets import GoogleSheetsAdapter

COLUMNS = [
    "vendor_name", "invoice_number", "invoice_date", "total",
    "gl_code", "anomaly_flags",
]


def make_invoice(
    vendor: str = "ACME Inc.",
    number: str = "INV-001",
    total: str = "1000.00",
    flags: list[str] | None = None,
) -> Invoice:
    from datetime import date
    inv = Invoice(source_file=Path("test.pdf"), template_used="_default")
    inv.vendor_name = vendor
    inv.invoice_number = number
    inv.total = Decimal(total)
    inv.invoice_date = date(2024, 1, 15)
    inv.gl_code = "6100"
    inv.anomaly_flags = flags or []
    return inv


# ── CSVAdapter ────────────────────────────────────────────────────────────────

class TestCSVAdapter:
    def test_creates_file_with_header(self, tmp_path: Path) -> None:
        path = tmp_path / "out.csv"
        adapter = CSVAdapter(file_path=path, columns=COLUMNS)
        adapter.write_rows([make_invoice()])
        assert path.exists()
        with open(path, newline="", encoding="utf-8") as fh:
            reader = csv.DictReader(fh)
            assert reader.fieldnames == COLUMNS

    def test_writes_correct_values(self, tmp_path: Path) -> None:
        path = tmp_path / "out.csv"
        adapter = CSVAdapter(file_path=path, columns=COLUMNS)
        adapter.write_rows([make_invoice(vendor="Widgets Ltd", number="W-99", total="500.00")])
        with open(path, newline="", encoding="utf-8") as fh:
            rows = list(csv.DictReader(fh))
        assert len(rows) == 1
        assert rows[0]["vendor_name"] == "Widgets Ltd"
        assert rows[0]["invoice_number"] == "W-99"
        assert rows[0]["total"] == "500.00"

    def test_returns_row_count(self, tmp_path: Path) -> None:
        path = tmp_path / "out.csv"
        adapter = CSVAdapter(file_path=path, columns=COLUMNS)
        n = adapter.write_rows([make_invoice(), make_invoice(number="INV-002")])
        assert n == 2

    def test_append_mode_adds_rows(self, tmp_path: Path) -> None:
        path = tmp_path / "out.csv"
        adapter = CSVAdapter(file_path=path, columns=COLUMNS, append=True)
        adapter.write_rows([make_invoice(number="INV-001")])
        adapter.write_rows([make_invoice(number="INV-002")])
        with open(path, newline="", encoding="utf-8") as fh:
            rows = list(csv.DictReader(fh))
        assert len(rows) == 2
        assert rows[0]["invoice_number"] == "INV-001"
        assert rows[1]["invoice_number"] == "INV-002"

    def test_append_mode_no_duplicate_header(self, tmp_path: Path) -> None:
        path = tmp_path / "out.csv"
        adapter = CSVAdapter(file_path=path, columns=COLUMNS, append=True)
        adapter.write_rows([make_invoice()])
        adapter.write_rows([make_invoice(number="INV-002")])
        with open(path, newline="", encoding="utf-8") as fh:
            lines = fh.readlines()
        header_lines = [ln for ln in lines if ln.startswith("vendor_name")]
        assert len(header_lines) == 1

    def test_overwrite_mode_replaces_file(self, tmp_path: Path) -> None:
        path = tmp_path / "out.csv"
        CSVAdapter(file_path=path, columns=COLUMNS, append=False).write_rows(
            [make_invoice(number="OLD")]
        )
        CSVAdapter(file_path=path, columns=COLUMNS, append=False).write_rows(
            [make_invoice(number="NEW")]
        )
        with open(path, newline="", encoding="utf-8") as fh:
            rows = list(csv.DictReader(fh))
        assert len(rows) == 1
        assert rows[0]["invoice_number"] == "NEW"

    def test_creates_parent_directories(self, tmp_path: Path) -> None:
        path = tmp_path / "nested" / "deep" / "out.csv"
        adapter = CSVAdapter(file_path=path, columns=COLUMNS)
        adapter.write_rows([make_invoice()])
        assert path.exists()

    def test_anomaly_flags_joined_with_semicolon(self, tmp_path: Path) -> None:
        path = tmp_path / "out.csv"
        adapter = CSVAdapter(file_path=path, columns=COLUMNS)
        adapter.write_rows([make_invoice(flags=["amount_threshold", "future_date"])])
        with open(path, newline="", encoding="utf-8") as fh:
            rows = list(csv.DictReader(fh))
        assert rows[0]["anomaly_flags"] == "amount_threshold; future_date"

    def test_empty_invoice_list_writes_only_header(self, tmp_path: Path) -> None:
        path = tmp_path / "out.csv"
        adapter = CSVAdapter(file_path=path, columns=COLUMNS, append=False)
        adapter.write_rows([])
        with open(path, newline="", encoding="utf-8") as fh:
            lines = [ln for ln in fh.readlines() if ln.strip()]
        assert len(lines) == 1  # header only

    def test_extra_invoice_fields_ignored(self, tmp_path: Path) -> None:
        """DictWriter with extrasaction='ignore' should not raise for unmapped fields."""
        path = tmp_path / "out.csv"
        adapter = CSVAdapter(file_path=path, columns=["vendor_name", "total"])
        adapter.write_rows([make_invoice()])  # invoice has many more fields
        with open(path, newline="", encoding="utf-8") as fh:
            rows = list(csv.DictReader(fh))
        assert list(rows[0].keys()) == ["vendor_name", "total"]


# ── ExcelAdapter ──────────────────────────────────────────────────────────────

class TestExcelAdapter:
    def test_creates_file(self, tmp_path: Path) -> None:
        import openpyxl
        path = tmp_path / "out.xlsx"
        adapter = ExcelAdapter(file_path=path, columns=COLUMNS)
        adapter.write_rows([make_invoice()])
        assert path.exists()
        wb = openpyxl.load_workbook(str(path))
        assert "Invoices" in wb.sheetnames

    def test_header_row_written(self, tmp_path: Path) -> None:
        import openpyxl
        path = tmp_path / "out.xlsx"
        adapter = ExcelAdapter(file_path=path, columns=COLUMNS)
        adapter.write_rows([make_invoice()])
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Invoices"]
        header = [ws.cell(1, c).value for c in range(1, len(COLUMNS) + 1)]
        assert header == COLUMNS

    def test_data_row_written(self, tmp_path: Path) -> None:
        import openpyxl
        path = tmp_path / "out.xlsx"
        adapter = ExcelAdapter(file_path=path, columns=COLUMNS)
        adapter.write_rows([make_invoice(vendor="TestCorp", number="T-1")])
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Invoices"]
        assert ws.cell(2, 1).value == "TestCorp"
        assert ws.cell(2, 2).value == "T-1"

    def test_returns_row_count(self, tmp_path: Path) -> None:
        path = tmp_path / "out.xlsx"
        adapter = ExcelAdapter(file_path=path, columns=COLUMNS)
        n = adapter.write_rows([make_invoice(), make_invoice(number="INV-002")])
        assert n == 2

    def test_append_mode_adds_rows(self, tmp_path: Path) -> None:
        import openpyxl
        path = tmp_path / "out.xlsx"
        adapter = ExcelAdapter(file_path=path, columns=COLUMNS, append=True)
        adapter.write_rows([make_invoice(number="INV-001")])
        adapter.write_rows([make_invoice(number="INV-002")])
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Invoices"]
        assert ws.max_row == 3  # header + 2 data rows

    def test_overwrite_mode_replaces_file(self, tmp_path: Path) -> None:
        import openpyxl
        path = tmp_path / "out.xlsx"
        ExcelAdapter(file_path=path, columns=COLUMNS, append=False).write_rows(
            [make_invoice(number="OLD")]
        )
        ExcelAdapter(file_path=path, columns=COLUMNS, append=False).write_rows(
            [make_invoice(number="NEW")]
        )
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Invoices"]
        assert ws.max_row == 2  # header + 1 data row
        assert ws.cell(2, 2).value == "NEW"

    def test_custom_sheet_name(self, tmp_path: Path) -> None:
        import openpyxl
        path = tmp_path / "out.xlsx"
        adapter = ExcelAdapter(file_path=path, columns=COLUMNS, sheet_name="Q1 2024")
        adapter.write_rows([make_invoice()])
        wb = openpyxl.load_workbook(str(path))
        assert "Q1 2024" in wb.sheetnames

    def test_creates_parent_directories(self, tmp_path: Path) -> None:
        path = tmp_path / "nested" / "deep" / "out.xlsx"
        adapter = ExcelAdapter(file_path=path, columns=COLUMNS)
        adapter.write_rows([make_invoice()])
        assert path.exists()


# ── GoogleSheetsAdapter ───────────────────────────────────────────────────────

class TestGoogleSheetsAdapter:
    def _make_adapter(self) -> GoogleSheetsAdapter:
        return GoogleSheetsAdapter(
            spreadsheet_id="fake-sheet-id",
            columns=COLUMNS,
            sheet_name="Invoices",
            credentials_env="GOOGLE_SHEETS_SERVICE_ACCOUNT",
        )

    def test_raises_if_env_var_missing(self, monkeypatch: pytest.MonkeyPatch) -> None:
        monkeypatch.delenv("GOOGLE_SHEETS_SERVICE_ACCOUNT", raising=False)
        adapter = self._make_adapter()
        with pytest.raises(RuntimeError, match="GOOGLE_SHEETS_SERVICE_ACCOUNT"):
            adapter._get_client()

    def test_write_rows_appends_to_existing_sheet(self, monkeypatch: pytest.MonkeyPatch) -> None:
        monkeypatch.setenv("GOOGLE_SHEETS_SERVICE_ACCOUNT", "/fake/creds.json")
        mock_ws = MagicMock()
        mock_spreadsheet = MagicMock()
        mock_spreadsheet.worksheet.return_value = mock_ws
        mock_client = MagicMock()
        mock_client.open_by_key.return_value = mock_spreadsheet

        adapter = self._make_adapter()
        with patch.object(adapter, "_get_client", return_value=mock_client):
            n = adapter.write_rows([make_invoice(vendor="Corp A"), make_invoice(vendor="Corp B")])

        assert n == 2
        assert mock_ws.append_row.call_count == 2

    def test_write_rows_creates_sheet_if_missing(self, monkeypatch: pytest.MonkeyPatch) -> None:
        import gspread
        monkeypatch.setenv("GOOGLE_SHEETS_SERVICE_ACCOUNT", "/fake/creds.json")
        mock_ws = MagicMock()
        mock_spreadsheet = MagicMock()
        mock_spreadsheet.worksheet.side_effect = gspread.exceptions.WorksheetNotFound
        mock_spreadsheet.add_worksheet.return_value = mock_ws
        mock_client = MagicMock()
        mock_client.open_by_key.return_value = mock_spreadsheet

        adapter = self._make_adapter()
        with patch.object(adapter, "_get_client", return_value=mock_client):
            adapter.write_rows([make_invoice()])

        mock_spreadsheet.add_worksheet.assert_called_once_with(
            title="Invoices", rows=1000, cols=20
        )
        mock_ws.append_row.assert_any_call(COLUMNS)

    def test_row_values_match_columns(self, monkeypatch: pytest.MonkeyPatch) -> None:
        monkeypatch.setenv("GOOGLE_SHEETS_SERVICE_ACCOUNT", "/fake/creds.json")
        mock_ws = MagicMock()
        mock_spreadsheet = MagicMock()
        mock_spreadsheet.worksheet.return_value = mock_ws
        mock_client = MagicMock()
        mock_client.open_by_key.return_value = mock_spreadsheet

        inv = make_invoice(vendor="ZZZ Corp", number="Z-99", total="250.00")
        adapter = self._make_adapter()
        with patch.object(adapter, "_get_client", return_value=mock_client):
            adapter.write_rows([inv])

        call_args = mock_ws.append_row.call_args[0][0]
        assert call_args[0] == "ZZZ Corp"
        assert call_args[1] == "Z-99"

    def test_get_client_builds_gspread_client(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """sheets.py:46-50 — _get_client loads creds and returns a gspread client."""
        monkeypatch.setenv("GOOGLE_SHEETS_SERVICE_ACCOUNT", "/fake/creds.json")

        mock_creds = MagicMock()
        mock_client = MagicMock()

        with (
            patch("google.oauth2.service_account.Credentials.from_service_account_file",
                  return_value=mock_creds),
            patch("gspread.authorize", return_value=mock_client),
        ):
            adapter = self._make_adapter()
            result = adapter._get_client()

        assert result is mock_client


# ── build_adapter factory ─────────────────────────────────────────────────────

class TestBuildAdapter:
    def test_csv_adapter(self, tmp_path: Path) -> None:
        cfg = {"adapter": "csv", "csv": {"file": str(tmp_path / "out.csv"), "append": True}}
        adapter = build_adapter(cfg, COLUMNS)
        assert isinstance(adapter, CSVAdapter)

    def test_excel_adapter(self, tmp_path: Path) -> None:
        cfg = {"adapter": "excel", "excel": {"file": str(tmp_path / "out.xlsx")}}
        adapter = build_adapter(cfg, COLUMNS)
        assert isinstance(adapter, ExcelAdapter)

    def test_google_sheets_adapter(self) -> None:
        cfg = {
            "adapter": "google_sheets",
            "google_sheets": {"spreadsheet_id": "abc123"},
        }
        adapter = build_adapter(cfg, COLUMNS)
        assert isinstance(adapter, GoogleSheetsAdapter)

    def test_default_adapter_is_csv(self, tmp_path: Path) -> None:
        adapter = build_adapter({}, COLUMNS)
        assert isinstance(adapter, CSVAdapter)

    def test_unknown_adapter_raises(self) -> None:
        with pytest.raises(ValueError, match="Unknown output adapter"):
            build_adapter({"adapter": "postgresql"}, COLUMNS)

    def test_google_sheets_missing_spreadsheet_id_raises(self) -> None:
        cfg = {"adapter": "google_sheets", "google_sheets": {}}
        with pytest.raises(KeyError):
            build_adapter(cfg, COLUMNS)
